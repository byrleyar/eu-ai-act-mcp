"""EU AI Act Compliance Report Generator.

Aggregates audit_results.json files produced by audit_processor.py into
batch summary reports: batch_report.csv (Plan 01) and batch_report.xlsx (Plan 02).
"""

import csv
import json
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import typer
from rich.console import Console

# ---------------------------------------------------------------------------
# Section map: 8 sections -> 80 question IDs
# Extracted from automated_testing/Hackathon Compliance Form Validation.xlsx
# ---------------------------------------------------------------------------

SECTION_MAP: dict[str, list[str]] = {
    "General Information": [
        "date_last_updated", "doc_version_number", "legal_name", "model_name",
        "model_authenticity", "release_date", "union_release_date", "model_dependencies",
    ],  # 8 questions
    "Model Properties": [
        "model_architecture", "design_specs",
        "input_modalities_text_check", "input_modalities_text_max",
        "input_modalities_images_check", "input_modalities_images_max",
        "input_modalities_audio_check", "input_modalities_audio_max",
        "input_modalities_video_check", "input_modalities_video_max",
        "input_modalities_other_check", "input_modalities_other_max",
        "output_modalities_text_check", "output_modalities_text_max",
        "output_modalities_images_check", "output_modalities_images_max",
        "output_modalities_audio_check", "output_modalities_audio_max",
        "output_modalities_video_check", "output_modalities_video_max",
        "output_modalities_other_check", "output_modalities_other_max",
        "total_model_size", "total_model_size_500m", "total_model_size_5b",
        "total_model_size_15b", "total_model_size_50b", "total_model_size_100b",
        "total_model_size_500b", "total_model_size_1t", "total_model_size_max",
    ],  # 31 questions
    "Methods of Distribution": [
        "distribution_channels_aio_nca", "distribution_channels_dps",
        "license_link", "license_category", "license_assets",
    ],  # 5 questions
    "Use": [
        "acceptable_use_policy", "intended_uses", "type_and_nature",
        "technical_means", "required_hardware", "required_software",
    ],  # 6 questions
    "Training Process": [
        "design_specifications", "decision_rationale",
    ],  # 2 questions
    "Information on Data": [
        "data_training_type_text", "data_training_type_images",
        "data_training_type_audio", "data_training_type_video",
        "data_training_type_other",
        "data_provenance_web", "data_provenance_private",
        "data_provenance_user", "data_provenance_public",
        "data_provenance_synthetic", "data_provenance_other_check",
        "data_provenance_other", "how_data_obtained",
        "data_points_ncas", "data_points_aio",
        "scope", "data_curation", "detect_unsuitability", "detect_biases",
    ],  # 19 questions
    "Computational Resources": [
        "training_time_ncas", "training_time_aio",
        "computation_used_ncas", "computation_used_aio",
        "computation_methodology",
    ],  # 5 questions
    "Energy Consumption": [
        "energy_used", "energy_methodology",
        "benchmark_computation", "computation_measurement_methodology",
    ],  # 4 questions
}
# Total: 80 questions -- matches questions.json exactly (verified)

# ---------------------------------------------------------------------------
# Startup assertion: SECTION_MAP must cover all IDs in questions.json exactly
# ---------------------------------------------------------------------------

_mapped_ids = {qid for qids in SECTION_MAP.values() for qid in qids}
_expected_ids = {
    q["id"]
    for q in json.load(open(Path(__file__).parent / "questions.json"))
}
assert _mapped_ids == _expected_ids, (
    f"SECTION_MAP drift: missing={_expected_ids - _mapped_ids}, "
    f"extra={_mapped_ids - _expected_ids}"
)

# ---------------------------------------------------------------------------
# CSV field names
# ---------------------------------------------------------------------------

CSV_FIELDNAMES = [
    "model_id",
    "status",
    "total_fields",
    "score_1",
    "score_1i",
    "score_2",
    "score_3",
    "score_4",
    "accuracy_pct",
    "hallucination_pct",
    "pass_pct",
    "fail_pct",
]

# ---------------------------------------------------------------------------
# Module-level console
# ---------------------------------------------------------------------------

console = Console()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def load_audit_results(model_dir: Path) -> dict | None:
    """Load audit_results.json from model_dir.

    Returns a dict with keys:
      - scores_by_qid: {question_id: score}
      - summary: raw summary dict from audit_results.json
      - model_id: str
      - audit_timestamp: str

    Returns None if audit_results.json is missing.
    """
    audit_path = model_dir / "audit_results.json"
    if not audit_path.exists():
        return None
    with open(audit_path, encoding="utf-8") as f:
        data = json.load(f)
    scores_by_qid = {fa["question_id"]: fa["score"] for fa in data["field_audits"]}
    return {
        "scores_by_qid": scores_by_qid,
        "summary": data.get("summary", {}),
        "model_id": data.get("model_id", model_dir.name),
        "audit_timestamp": data.get("audit_timestamp", ""),
    }


def compute_metrics(score_counts: dict) -> dict:
    """Compute compliance audit metrics from score counts.

    Args:
        score_counts: dict with keys score_1, score_1i, score_2, score_3,
                      score_4, total_fields.

    Returns:
        dict with keys: accuracy_rate, hallucination_rate, gap_rate, na_rate,
        pass_rate, fail_rate. All values are floats in [0.0, 1.0].
    """
    total = score_counts.get("total_fields", 0)
    if total == 0:
        return {
            "accuracy_rate": 0.0,
            "hallucination_rate": 0.0,
            "gap_rate": 0.0,
            "na_rate": 0.0,
            "pass_rate": 0.0,
            "fail_rate": 0.0,
        }

    s1 = score_counts.get("score_1", 0)
    s1i = score_counts.get("score_1i", 0)
    s2 = score_counts.get("score_2", 0)
    s3 = score_counts.get("score_3", 0)
    s4 = score_counts.get("score_4", 0)

    return {
        "accuracy_rate": (s1 + s1i) / total,
        "hallucination_rate": s2 / total,
        "gap_rate": s3 / total,
        "na_rate": s4 / total,
        "pass_rate": (s1 + s1i + s4) / total,
        "fail_rate": (s2 + s3) / total,
    }


# ---------------------------------------------------------------------------
# ReportGenerator class
# ---------------------------------------------------------------------------


class ReportGenerator:
    """Aggregates audit_results.json files into batch summary reports."""

    def discover_audited_models(self, batch_dir: Path) -> list[dict]:
        """Iterate sorted subdirectories of batch_dir and collect audit data.

        For each subdirectory that contains audit_results.json, load the
        results and compute metrics. Summary keys (score_1_count, etc.) are
        remapped to compute_metrics keys (score_1, etc.) before calling
        compute_metrics.

        Returns:
            List of dicts with keys:
              model_id, model_dir, scores_by_qid, summary, metrics
        """
        models = []
        for child in sorted(batch_dir.iterdir()):
            if not child.is_dir():
                continue
            result = load_audit_results(child)
            if result is None:
                continue

            summary = result["summary"]
            # Remap _count suffixed keys to plain keys expected by compute_metrics
            score_counts = {
                "score_1": summary.get("score_1_count", 0),
                "score_1i": summary.get("score_1i_count", 0),
                "score_2": summary.get("score_2_count", 0),
                "score_3": summary.get("score_3_count", 0),
                "score_4": summary.get("score_4_count", 0),
                "total_fields": summary.get("total_fields", 0),
            }
            metrics = compute_metrics(score_counts)

            models.append({
                "model_id": result["model_id"],
                "model_dir": child,
                "scores_by_qid": result["scores_by_qid"],
                "summary": summary,
                "metrics": metrics,
                "score_counts": score_counts,
            })

        return models

    def write_csv(self, batch_dir: Path, models: list[dict]) -> Path:
        """Write batch_report.csv to batch_dir.

        Each row contains per-model score counts and percentage metrics.
        Percentages are 0-100 floats (rate * 100, rounded to 2 decimals).

        Returns:
            Path to the written CSV file.
        """
        csv_path = batch_dir / "batch_report.csv"
        with open(csv_path, mode="w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=CSV_FIELDNAMES)
            writer.writeheader()
            for model in models:
                sc = model["score_counts"]
                m = model["metrics"]
                writer.writerow({
                    "model_id": model["model_id"],
                    "status": "audited",
                    "total_fields": sc["total_fields"],
                    "score_1": sc["score_1"],
                    "score_1i": sc["score_1i"],
                    "score_2": sc["score_2"],
                    "score_3": sc["score_3"],
                    "score_4": sc["score_4"],
                    "accuracy_pct": round(m["accuracy_rate"] * 100, 2),
                    "hallucination_pct": round(m["hallucination_rate"] * 100, 2),
                    "pass_pct": round(m["pass_rate"] * 100, 2),
                    "fail_pct": round(m["fail_rate"] * 100, 2),
                })
        return csv_path

    def write_xlsx(self, batch_dir: Path, models: list[dict]) -> Path:
        """Write batch_report.xlsx to batch_dir with 3 sheets.

        Sheet 1 -- Per-Field Detail: one row per question with scores and sidebar aggregate.
        Sheet 2 -- Section Summary: N/A count per section per model.
        Sheet 3 -- Executive Summary: metadata, metrics, ranking, flagged issues.

        Returns:
            Path to the written xlsx file.
        """
        wb = openpyxl.Workbook()

        # --- Shared style helpers ---
        def hdr_style(cell: openpyxl.cell.Cell) -> None:
            """Apply blue header style to a cell."""
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _bold_sz(cell: openpyxl.cell.Cell, size: int = 11) -> None:
            cell.font = Font(bold=True, size=size)

        GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
        ORANGE_FILL = PatternFill("solid", fgColor="FFC7CE")
        GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")

        score_fill_map = {
            "1": GREEN_FILL,
            "1i": GREEN_FILL,
            "2": ORANGE_FILL,
            "3": ORANGE_FILL,
            "4": GRAY_FILL,
        }

        # -----------------------------------------------------------------
        # Sheet 1: Per-Field Detail
        # -----------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Per-Field Detail"

        model_count = len(models)
        # Column indices: A=1 (Section), B=2 (Question ID), then model columns start at 3
        model_col_start = 3
        # Sidebar starts after model columns + 1 gap
        sidebar_col_start = model_col_start + model_count + 1

        # Header row
        ws1.cell(row=1, column=1, value="Section")
        ws1.cell(row=1, column=2, value="Question ID")
        for i, model in enumerate(models):
            c = ws1.cell(row=1, column=model_col_start + i, value=model["model_id"])
            hdr_style(c)
        # Sidebar header
        sidebar_headers = [
            "Model", "Direct (1)", "%(1)", "Inferred (1i)", "%(1i)",
            "Hallucinated (2)", "%(2)", "Incomplete (3)", "%(3)", "N/A (4)", "%(4)",
        ]
        for j, sh in enumerate(sidebar_headers):
            c = ws1.cell(row=1, column=sidebar_col_start + j, value=sh)
            hdr_style(c)
        # Style first two header cells
        for col in (1, 2):
            hdr_style(ws1.cell(row=1, column=col))

        # Data rows
        data_row = 2
        all_questions_ordered = []
        for section, qids in SECTION_MAP.items():
            for qid in qids:
                all_questions_ordered.append((section, qid))

        for section, qid in all_questions_ordered:
            ws1.cell(row=data_row, column=1, value=section)
            ws1.cell(row=data_row, column=2, value=qid)
            for i, model in enumerate(models):
                score_val = model["scores_by_qid"].get(qid, "")
                c = ws1.cell(row=data_row, column=model_col_start + i, value=score_val)
                fill = score_fill_map.get(str(score_val))
                if fill:
                    c.fill = fill
                    c.alignment = Alignment(horizontal="center")
            data_row += 1

        # Sidebar aggregate block (data rows 2+)
        for i, model in enumerate(models):
            sc = model["score_counts"]
            m = model["metrics"]
            total = sc["total_fields"]
            s1 = sc["score_1"]
            s1i = sc["score_1i"]
            s2 = sc["score_2"]
            s3 = sc["score_3"]
            s4 = sc["score_4"]
            sidebar_row = 2 + i
            ws1.cell(row=sidebar_row, column=sidebar_col_start, value=model["model_id"])
            ws1.cell(row=sidebar_row, column=sidebar_col_start + 1, value=s1)
            pct1 = ws1.cell(row=sidebar_row, column=sidebar_col_start + 2, value=s1 / total if total else 0.0)
            pct1.number_format = "0.0%"
            ws1.cell(row=sidebar_row, column=sidebar_col_start + 3, value=s1i)
            pct1i = ws1.cell(row=sidebar_row, column=sidebar_col_start + 4, value=s1i / total if total else 0.0)
            pct1i.number_format = "0.0%"
            ws1.cell(row=sidebar_row, column=sidebar_col_start + 5, value=s2)
            pct2 = ws1.cell(row=sidebar_row, column=sidebar_col_start + 6, value=m["hallucination_rate"])
            pct2.number_format = "0.0%"
            ws1.cell(row=sidebar_row, column=sidebar_col_start + 7, value=s3)
            pct3 = ws1.cell(row=sidebar_row, column=sidebar_col_start + 8, value=m["gap_rate"])
            pct3.number_format = "0.0%"
            ws1.cell(row=sidebar_row, column=sidebar_col_start + 9, value=s4)
            pct4 = ws1.cell(row=sidebar_row, column=sidebar_col_start + 10, value=m["na_rate"])
            pct4.number_format = "0.0%"

        # Average row in sidebar
        if model_count > 0:
            avg_row = 2 + model_count
            ws1.cell(row=avg_row, column=sidebar_col_start, value="Average")
            _bold_sz(ws1.cell(row=avg_row, column=sidebar_col_start))
            # Only populate percentage columns
            pct_col_offsets = [2, 4, 6, 8, 10]
            for offset in pct_col_offsets:
                vals = [
                    ws1.cell(row=2 + i, column=sidebar_col_start + offset).value
                    for i in range(model_count)
                ]
                avg_val = sum(v for v in vals if v is not None) / model_count
                c = ws1.cell(row=avg_row, column=sidebar_col_start + offset, value=avg_val)
                c.number_format = "0.0%"

        # Column widths
        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 35
        for i in range(model_count):
            col_letter = openpyxl.utils.get_column_letter(model_col_start + i)
            ws1.column_dimensions[col_letter].width = 10
        # Sidebar widths
        sidebar_widths = [30, 10, 8, 14, 8, 16, 8, 14, 8, 10, 8]
        for j, w in enumerate(sidebar_widths):
            col_letter = openpyxl.utils.get_column_letter(sidebar_col_start + j)
            ws1.column_dimensions[col_letter].width = w

        # Freeze pane at A2
        ws1.freeze_panes = "A2"

        # Auto-filter on data columns A through last model column
        last_model_col = openpyxl.utils.get_column_letter(model_col_start + model_count - 1) if model_count else "B"
        ws1.auto_filter.ref = f"A1:{last_model_col}1"

        # -----------------------------------------------------------------
        # Sheet 2: Section Summary
        # -----------------------------------------------------------------
        ws2 = wb.create_sheet("Section Summary")

        # Header
        ws2.cell(row=1, column=1, value="Section")
        hdr_style(ws2.cell(row=1, column=1))
        for i, model in enumerate(models):
            c = ws2.cell(row=1, column=2 + i, value=model["model_id"])
            hdr_style(c)
        total_4s_col = 2 + model_count
        avg_4s_col = total_4s_col + 1
        qs_in_section_col = avg_4s_col + 1
        for col, header in [
            (total_4s_col, "Total 4s (N/A)"),
            (avg_4s_col, "Avg 4s"),
            (qs_in_section_col, "Questions in Section"),
        ]:
            c = ws2.cell(row=1, column=col, value=header)
            hdr_style(c)

        # Section rows
        for row_idx, (section, qids) in enumerate(SECTION_MAP.items(), start=2):
            ws2.cell(row=row_idx, column=1, value=section)
            model_counts_4 = []
            for i, model in enumerate(models):
                count_4 = sum(
                    1 for qid in qids
                    if model["scores_by_qid"].get(qid) == "4"
                )
                ws2.cell(row=row_idx, column=2 + i, value=count_4)
                model_counts_4.append(count_4)
            total_4 = sum(model_counts_4)
            avg_4 = total_4 / model_count if model_count else 0.0
            ws2.cell(row=row_idx, column=total_4s_col, value=total_4)
            ws2.cell(row=row_idx, column=avg_4s_col, value=round(avg_4, 2))
            ws2.cell(row=row_idx, column=qs_in_section_col, value=len(qids))

        # Total row
        total_row = 2 + len(SECTION_MAP)
        ws2.cell(row=total_row, column=1, value="Total")
        _bold_sz(ws2.cell(row=total_row, column=1))
        for i in range(model_count):
            col = 2 + i
            col_sum = sum(
                ws2.cell(row=2 + r, column=col).value or 0
                for r in range(len(SECTION_MAP))
            )
            c = ws2.cell(row=total_row, column=col, value=col_sum)
            _bold_sz(c)
        # Total 4s total
        grand_total_4s = sum(
            ws2.cell(row=2 + r, column=total_4s_col).value or 0
            for r in range(len(SECTION_MAP))
        )
        c = ws2.cell(row=total_row, column=total_4s_col, value=grand_total_4s)
        _bold_sz(c)
        overall_avg_4s = grand_total_4s / model_count if model_count else 0.0
        c = ws2.cell(row=total_row, column=avg_4s_col, value=round(overall_avg_4s, 2))
        _bold_sz(c)
        total_questions = sum(len(qids) for qids in SECTION_MAP.values())
        c = ws2.cell(row=total_row, column=qs_in_section_col, value=total_questions)
        _bold_sz(c)

        # Column widths for Sheet 2
        ws2.column_dimensions["A"].width = 25
        for i in range(model_count):
            col_letter = openpyxl.utils.get_column_letter(2 + i)
            ws2.column_dimensions[col_letter].width = 12
        for col in (total_4s_col, avg_4s_col, qs_in_section_col):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

        # -----------------------------------------------------------------
        # Sheet 3: Executive Summary
        # -----------------------------------------------------------------
        ws3 = wb.create_sheet("Executive Summary")

        # Row 1: Title
        title_cell = ws3.cell(row=1, column=1, value="EU AI Act Compliance Audit - Executive Summary")
        title_cell.font = Font(bold=True, size=14)

        # Metadata
        ws3.cell(row=3, column=1, value="Batch:")
        ws3.cell(row=3, column=2, value=batch_dir.name)
        ws3.cell(row=4, column=1, value="Report Date:")
        ws3.cell(row=4, column=2, value=date.today().isoformat())
        ws3.cell(row=5, column=1, value="Models Audited:")
        ws3.cell(row=5, column=2, value=model_count)

        # Overall Metrics header
        om_cell = ws3.cell(row=7, column=1, value="Overall Metrics")
        om_cell.font = Font(bold=True, size=12)

        # Compute averages across models
        def _avg_metric(key: str) -> float:
            if not models:
                return 0.0
            return sum(m["metrics"][key] for m in models) / len(models)

        overall_metrics = [
            ("Accuracy Rate", _avg_metric("accuracy_rate")),
            ("Hallucination Rate", _avg_metric("hallucination_rate")),
            ("Gap Rate (Incomplete)", _avg_metric("gap_rate")),
            ("N/A Rate", _avg_metric("na_rate")),
            ("Pass Rate", _avg_metric("pass_rate")),
            ("Fail Rate", _avg_metric("fail_rate")),
        ]

        # Metric label/value header
        ws3.cell(row=8, column=1, value="Metric").font = Font(bold=True)
        ws3.cell(row=8, column=2, value="Value").font = Font(bold=True)
        for offset, (metric_name, metric_val) in enumerate(overall_metrics, start=9):
            ws3.cell(row=offset, column=1, value=metric_name)
            c = ws3.cell(row=offset, column=2, value=metric_val)
            c.number_format = "0.0%"

        # Per-model ranking
        ranking_title_row = 15
        pm_cell = ws3.cell(row=ranking_title_row, column=1, value="Per-Model Ranking (by Accuracy)")
        pm_cell.font = Font(bold=True, size=12)

        rank_hdr_row = ranking_title_row + 1
        for col, hdr in enumerate(["Rank", "Model", "Accuracy", "Hallucination", "Pass", "Fail"], start=1):
            c = ws3.cell(row=rank_hdr_row, column=col, value=hdr)
            c.font = Font(bold=True)

        sorted_models = sorted(models, key=lambda m: m["metrics"]["accuracy_rate"], reverse=True)
        for rank_idx, model in enumerate(sorted_models, start=1):
            r = rank_hdr_row + rank_idx
            ws3.cell(row=r, column=1, value=rank_idx)
            ws3.cell(row=r, column=2, value=model["model_id"])
            m = model["metrics"]
            for col_offset, key in enumerate(["accuracy_rate", "hallucination_rate", "pass_rate", "fail_rate"], start=3):
                c = ws3.cell(row=r, column=col_offset, value=m[key])
                c.number_format = "0.0%"

        # Flagged issues section
        flagged_start_row = rank_hdr_row + len(models) + 3
        fi_cell = ws3.cell(row=flagged_start_row, column=1, value="Flagged Issues")
        fi_cell.font = Font(bold=True, size=12)

        flagged_rows = []
        for model in models:
            for qid, score in model["scores_by_qid"].items():
                if score == "2":
                    flagged_rows.append(
                        f"Model: {model['model_id']}, Field: {qid}, Score: 2 (Hallucinated)"
                    )
                elif score == "3":
                    flagged_rows.append(
                        f"Model: {model['model_id']}, Field: {qid}, Score: 3 (Incomplete)"
                    )

        if flagged_rows:
            for fi_idx, issue_text in enumerate(flagged_rows, start=1):
                ws3.cell(row=flagged_start_row + fi_idx, column=1, value=issue_text)
        else:
            ws3.cell(
                row=flagged_start_row + 1, column=1,
                value="No hallucinated or incomplete fields detected."
            )

        # Column widths for Sheet 3
        ws3.column_dimensions["A"].width = 35
        ws3.column_dimensions["B"].width = 30

        # -----------------------------------------------------------------
        # Save workbook
        # -----------------------------------------------------------------
        xlsx_path = batch_dir / "batch_report.xlsx"
        wb.save(xlsx_path)
        return xlsx_path

    def write_executive_summary_txt(self, batch_dir: Path, models: list[dict]) -> Path:
        """Write executive_summary.txt to batch_dir as readable plain text.

        Contains the same information as Sheet 3 of the xlsx (batch metadata,
        overall metrics, per-model ranking, flagged issues).

        Returns:
            Path to the written text file.
        """
        def _pct(rate: float) -> str:
            return f"{rate * 100:.1f}%"

        def _avg_metric(key: str) -> float:
            if not models:
                return 0.0
            return sum(m["metrics"][key] for m in models) / len(models)

        lines = [
            "EU AI Act Compliance Audit - Executive Summary",
            "================================================",
            f"Batch: {batch_dir.name}",
            f"Report Date: {date.today().isoformat()}",
            f"Models Audited: {len(models)}",
            "",
            "Overall Metrics",
            "---------------",
            f"Accuracy Rate:      {_pct(_avg_metric('accuracy_rate'))}",
            f"Hallucination Rate: {_pct(_avg_metric('hallucination_rate'))}",
            f"Gap Rate:           {_pct(_avg_metric('gap_rate'))}",
            f"N/A Rate:           {_pct(_avg_metric('na_rate'))}",
            f"Pass Rate:          {_pct(_avg_metric('pass_rate'))}",
            f"Fail Rate:          {_pct(_avg_metric('fail_rate'))}",
            "",
            "Per-Model Ranking (by Accuracy)",
            "-------------------------------",
        ]

        sorted_models = sorted(models, key=lambda m: m["metrics"]["accuracy_rate"], reverse=True)
        for rank_idx, model in enumerate(sorted_models, start=1):
            m = model["metrics"]
            lines.append(
                f"{rank_idx}. {model['model_id']} -- "
                f"Accuracy: {_pct(m['accuracy_rate'])}, "
                f"Pass: {_pct(m['pass_rate'])}, "
                f"Fail: {_pct(m['fail_rate'])}"
            )

        lines += ["", "Flagged Issues", "--------------"]
        flagged = []
        for model in models:
            for qid, score in model["scores_by_qid"].items():
                if score == "2":
                    flagged.append(f"[HALLUCINATED] {model['model_id']} / {qid}")
                elif score == "3":
                    flagged.append(f"[INCOMPLETE] {model['model_id']} / {qid}")
        if flagged:
            lines.extend(flagged)
        else:
            lines.append("No hallucinated or incomplete fields detected.")

        txt_path = batch_dir / "executive_summary.txt"
        txt_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return txt_path

    def generate(self, batch_dir: Path) -> None:
        """Orchestrate report generation for a batch directory.

        1. Discover all model subdirectories with audit_results.json
        2. Validate at least one model was audited
        3. Write batch_report.csv
        4. Write batch_report.xlsx
        5. Write executive_summary.txt
        6. Print summary to console
        """
        if not batch_dir.exists():
            console.print(
                f"[red]Error: batch directory '{batch_dir}' does not exist.[/red]"
            )
            raise typer.Exit(1)

        console.print(f"[bold green]Generating report for: {batch_dir}[/bold green]")

        models = self.discover_audited_models(batch_dir)

        if not models:
            console.print(
                "[red]No audit_results.json found. Run audit_processor.py first.[/red]"
            )
            raise typer.Exit(1)

        console.print(f"Models with audit data: {len(models)}")

        # Write CSV
        csv_path = self.write_csv(batch_dir, models)
        console.print(f"CSV report:       [cyan]{csv_path}[/cyan]")

        # Write xlsx
        xlsx_path = self.write_xlsx(batch_dir, models)
        console.print(f"Excel report:     [cyan]{xlsx_path}[/cyan]")

        # Write executive_summary.txt
        txt_path = self.write_executive_summary_txt(batch_dir, models)
        console.print(f"Executive summary:[cyan]{txt_path}[/cyan]")

        # Print console summary table
        console.print("\n[bold]Audit Report Summary[/bold]")
        console.print(
            f"{'Model':<45} {'Fields':>6} {'1':>4} {'1i':>4} "
            f"{'2':>4} {'3':>4} {'4':>4} {'Acc%':>6} {'Pass%':>6}"
        )
        console.print("-" * 90)
        for model in models:
            sc = model["score_counts"]
            m = model["metrics"]
            console.print(
                f"{model['model_id']:<45} {sc['total_fields']:>6} "
                f"{sc['score_1']:>4} {sc['score_1i']:>4} {sc['score_2']:>4} "
                f"{sc['score_3']:>4} {sc['score_4']:>4} "
                f"{m['accuracy_rate']*100:>6.1f} {m['pass_rate']*100:>6.1f}"
            )

        console.print(f"\n[bold green]Report complete![/bold green]")
        console.print(f"Models processed: {len(models)}")


# ---------------------------------------------------------------------------
# Typer CLI
# ---------------------------------------------------------------------------

app = typer.Typer(help="EU AI Act Compliance Report Generator")


@app.command()
def run(
    batch_dir: Path = typer.Argument(..., help="Path to batch directory"),
) -> None:
    """Generate aggregate compliance audit report for a batch directory."""
    generator = ReportGenerator()
    generator.generate(batch_dir)


if __name__ == "__main__":
    app()
